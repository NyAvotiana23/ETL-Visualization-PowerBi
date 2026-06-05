"""
╔══════════════════════════════════════════════════════════════════╗
║         ETL DATAWAREHOUSE — TP Modélisation DWH                 ║
║  Sources : MySQL SQL | TXT | Excel | JSON                        ║
║  Destination : PostgreSQL (schéma en étoile)                     ║
╚══════════════════════════════════════════════════════════════════╝

Architecture de la table de faits :
  FAIT_ANALYSE_REPRESENTANT
    ← DIM_TEMPS
    ← DIM_VENDEUR
    ← DIM_CLIENT
    ← DIM_PRODUIT
    ← DIM_GEO

Usage :
  python3 etl_pipeline.py [--host HOST] [--port PORT] [--db DB] [--user USER] [--password PWD]

Par défaut : localhost:5432, db=dwh_imprimantes, user=postgres
"""

import json
import re
import sys
import argparse
from datetime import datetime, date
from pathlib import Path

# ─── Config chemins ───────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent / "Dtw1"
SQL_FILE   = BASE_DIR / "01_mysql_ventes_rh.sql"
TXT_FILE   = BASE_DIR / "02_clients.txt"
EXCEL_FILE = BASE_DIR / "03_produits.xlsx"
JSON_FILE  = BASE_DIR / "04_feuilles_route.json"

# ─── Couleurs terminal ────────────────────────────────────────────────────────
G = "\033[92m"; Y = "\033[93m"; R = "\033[91m"; B = "\033[94m"; E = "\033[0m"
def ok(msg):   print(f"  {G}✓{E} {msg}")
def info(msg): print(f"  {B}→{E} {msg}")
def warn(msg): print(f"  {Y}⚠{E} {msg}")
def err(msg):  print(f"  {R}✗{E} {msg}")

# ══════════════════════════════════════════════════════════════════════════════
# ÉTAPE 1 — EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════
def extract_mysql_sql(path: Path) -> tuple[list, list]:
    """Parse le fichier SQL MySQL pour extraire vendeurs et ventes."""
    vendeurs, ventes = [], []
    content = path.read_text(encoding="cp1252")
    
    # Extraction vendeurs
    pattern_v = re.compile(
        r"INSERT INTO vendeurs VALUES \((\d+),'([^']+)','([^']+)',([0-9.]+),'([^']+)','([^']+)','([^']+)'\);"
    )
    for m in pattern_v.finditer(content):
        vendeurs.append({
            "id_vendeur":    int(m.group(1)),
            "nom":           m.group(2),
            "prenom":        m.group(3),
            "salaire":       float(m.group(4)),
            "date_embauche": m.group(5),
            "ville_base":    m.group(6),
            "groupe":        m.group(7),
        })
    
    # Extraction ventes
    pattern_vt = re.compile(
        r"INSERT INTO ventes VALUES \((\d+),(\d+),(\d+),(\d+),"
        r"'([^']+)','([^']+)',(\d+),([0-9.]+),([0-9.]+),"
        r"(\d+),([0-9.]+),([0-9.]+),(\d+),'([^']+)'\);"
    )
    for m in pattern_vt.finditer(content):
        ventes.append({
            "id_vente":        int(m.group(1)),
            "id_vendeur":      int(m.group(2)),
            "id_client":       int(m.group(3)),
            "id_produit":      int(m.group(4)),
            "date_commande":   m.group(5),
            "date_precommande":m.group(6),
            "quantite":        int(m.group(7)),
            "montant_vente":   float(m.group(8)),
            "montant_precom":  float(m.group(9)),
            "km_parcourus":    int(m.group(10)),
            "litres_essence":  float(m.group(11)),
            "frais_voyage":    float(m.group(12)),
            "nb_visites":      int(m.group(13)),
            "ville_visite":    m.group(14),
        })
    return vendeurs, ventes

def extract_txt(path: Path) -> list:
    """Parse le fichier TXT clients (séparateur |)."""
    clients = []
    lines = path.read_text(encoding="cp1252").splitlines()
    headers = None
    for line in lines:
        if line.startswith("#"):
            continue
        parts = line.strip().split("|")
        if headers is None:
            headers = parts
            continue
        if len(parts) == len(headers):
            clients.append(dict(zip(headers, parts)))
    return clients

def extract_excel(path: Path) -> list:
    """Lit le fichier Excel produits avec openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Produits"]
    produits = []
    headers = [ws.cell(1, col).value for col in range(1, 10)]
    for row in range(2, ws.max_row + 1):
        vals = [ws.cell(row, col).value for col in range(1, 10)]
        if vals[0] is None:
            continue
        produits.append(dict(zip(headers, vals)))
    return produits

def extract_json(path: Path) -> list:
    """Charge les feuilles de route JSON."""
    return json.loads(path.read_text(encoding="utf-8"))

# ══════════════════════════════════════════════════════════════════════════════
# ÉTAPE 2 — TRANSFORMATION
# ══════════════════════════════════════════════════════════════════════════════
PROVINCES_MAP = {
    "Paris":"Île-de-France","Lyon":"Auvergne-Rhône-Alpes",
    "Marseille":"Provence-Alpes-Côte d''Azur","Toulouse":"Occitanie",
    "Nice":"Provence-Alpes-Côte d''Azur","Nantes":"Pays de la Loire",
    "Strasbourg":"Grand Est","Bordeaux":"Nouvelle-Aquitaine",
    "Lille":"Hauts-de-France","Rennes":"Bretagne",
    "Montpellier":"Occitanie","Grenoble":"Auvergne-Rhône-Alpes",
    "Toulon":"Provence-Alpes-Côte d''Azur","Dijon":"Bourgogne-Franche-Comté",
    "Angers":"Pays de la Loire","Nîmes":"Occitanie",
    "Saint-Etienne":"Auvergne-Rhône-Alpes","Clermont-Ferrand":"Auvergne-Rhône-Alpes",
    "Le Havre":"Normandie","Reims":"Grand Est",
}

def transform_dim_temps(ventes: list, feuilles: list) -> list:
    """
    Construit DIM_TEMPS à partir de toutes les dates présentes.
    Transformations : extraction année/mois/jour/semaine/trimestre/libellé.
    """
    dates_set = set()
    for v in ventes:
        dates_set.add(v["date_commande"])
        dates_set.add(v["date_precommande"])
    for f in feuilles:
        dates_set.add(f["date_deplacement"])
    
    dim_temps = []
    for i, d_str in enumerate(sorted(dates_set), 1):
        d = datetime.strptime(d_str, "%Y-%m-%d").date()
        dim_temps.append({
            "id_dim_temps":   i,
            "date_complete":  d_str,
            "annee":          d.year,
            "trimestre":      (d.month - 1) // 3 + 1,
            "mois":           d.month,
            "lib_mois":       ["Janvier","Février","Mars","Avril","Mai","Juin",
                               "Juillet","Août","Septembre","Octobre","Novembre","Décembre"][d.month-1],
            "semaine":        d.isocalendar()[1],
            "jour":           d.day,
            "lib_jour":       ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"][d.weekday()],
            "est_weekend":    d.weekday() >= 5,
        })
    return dim_temps

def transform_dim_vendeur(vendeurs: list) -> list:
    """
    Construit DIM_VENDEUR.
    Transformation : normalisation du groupe, ajout clé surrogate.
    """
    dim = []
    for i, v in enumerate(vendeurs, 1):
        groupe_norm = v["groupe"].strip().title()
        dim.append({
            "id_dim_vendeur":  i,
            "id_source":       v["id_vendeur"],
            "nom":             v["nom"].strip().upper(),
            "prenom":          v["prenom"].strip().title(),
            "nom_complet":     f"{v['prenom'].strip().title()} {v['nom'].strip().upper()}",
            "salaire":         float(v["salaire"]),
            "date_embauche":   v["date_embauche"],
            "ville_base":      v["ville_base"].strip(),
            "groupe":          groupe_norm,
        })
    return dim

def transform_dim_client(clients: list) -> list:
    """
    Construit DIM_CLIENT depuis le fichier TXT.
    Transformations : nettoyage, normalisation, ajout personne_ressource.
    """
    dim = []
    for i, c in enumerate(clients, 1):
        dim.append({
            "id_dim_client":      i,
            "id_source":          int(c["id_client"]),
            "nom":                c["nom"].strip().upper(),
            "prenom":             c["prenom"].strip().title(),
            "adresse":            c["adresse"].strip(),
            "ville":              c["ville"].strip(),
            "province":           c["province"].strip(),
            "pays":               c["pays"].strip(),
            "personne_ressource": c["personne_ressource"].strip(),
            "telephone":          c["telephone"].strip(),
            "email":              c["email"].strip().lower(),
            "segment":            c["segment"].strip(),
        })
    return dim

def transform_dim_produit(produits: list) -> list:
    """
    Construit DIM_PRODUIT depuis Excel.
    Transformations : conversion types, nettoyage valeurs None.
    """
    dim = []
    for p in produits:
        dim.append({
            "id_dim_produit":  int(p["id_produit"]),
            "nom":             str(p["nom_produit"]).strip(),
            "categorie":       str(p["categorie"]).strip(),
            "groupe":          str(p["groupe"]).strip(),
            "prix_unitaire":   float(p["prix_unitaire"]) if p["prix_unitaire"] else 0.0,
            "stock":           int(p["stock_disponible"]) if p["stock_disponible"] else 0,
            "fournisseur":     str(p["fournisseur"]).strip() if p["fournisseur"] else "Inconnu",
            "date_lancement":  str(p["date_lancement"]) if p["date_lancement"] else None,
            "actif":           str(p["actif"]).strip() == "Oui",
        })
    return dim

def transform_dim_geo(ventes: list, feuilles: list) -> list:
    """
    Construit DIM_GEO en collectant toutes les villes.
    Transformation : ajout province via mapping.
    """
    villes_set = set()
    for v in ventes:
        villes_set.add(v["ville_visite"])
    for f in feuilles:
        for ville in f.get("villes_etapes", []):
            villes_set.add(ville)
    
    dim = []
    for i, ville in enumerate(sorted(villes_set), 1):
        dim.append({
            "id_dim_geo": i,
            "ville":      ville,
            "province":   PROVINCES_MAP.get(ville, "Inconnue"),
            "pays":       "France",
        })
    return dim

def transform_fait(ventes: list, feuilles: list,
                   dim_temps: list, dim_vendeur: list,
                   dim_client: list, dim_produit: list, dim_geo: list) -> list:
    """
    Construit la table de faits en agrégeant ventes + feuilles de route.
    Clés surrogates résolues. Mesures calculées.
    """
    # Index rapides
    temps_idx   = {r["date_complete"]: r["id_dim_temps"]   for r in dim_temps}
    vendeur_idx = {r["id_source"]:     r["id_dim_vendeur"] for r in dim_vendeur}
    client_idx  = {r["id_source"]:     r["id_dim_client"]  for r in dim_client}
    produit_idx = {r["id_dim_produit"]:r["id_dim_produit"] for r in dim_produit}
    geo_idx     = {r["ville"]:         r["id_dim_geo"]     for r in dim_geo}

    # Index feuilles de route par (vendeur, date)
    feuilles_idx = {}
    for f in feuilles:
        key = (f["id_vendeur"], f["date_deplacement"])
        if key not in feuilles_idx:
            feuilles_idx[key] = []
        feuilles_idx[key].append(f)

    faits = []
    for v in ventes:
        id_t = temps_idx.get(v["date_commande"])
        id_v = vendeur_idx.get(v["id_vendeur"])
        id_c = client_idx.get(v["id_client"])
        id_p = produit_idx.get(v["id_produit"])
        id_g = geo_idx.get(v["ville_visite"])

        if not all([id_t, id_v, id_c, id_p, id_g]):
            continue  # Rejet si clé manquante (orphelin)

        # Enrichissement depuis feuilles de route
        frais_total = v["frais_voyage"]
        key = (v["id_vendeur"], v["date_commande"])
        if key in feuilles_idx:
            feuille = feuilles_idx[key][0]
            frais_total += feuille["frais"]["total"]

        # Calcul marge estimée (transformation métier)
        dim_p = next(p for p in dim_produit if p["id_dim_produit"] == id_p)
        cout_estime = dim_p["prix_unitaire"] * v["quantite"] * 0.60
        marge_estimee = round(v["montant_vente"] - cout_estime, 2)
        rentabilite = round(marge_estimee - frais_total, 2)

        faits.append({
            "id_dim_temps":    id_t,
            "id_dim_vendeur":  id_v,
            "id_dim_client":   id_c,
            "id_dim_geo":      id_g,
            "id_dim_produit":  id_p,
            "quantite_vendue": v["quantite"],
            "montant_vente":   v["montant_vente"],
            "montant_precom":  v["montant_precom"],
            "km_parcourus":    v["km_parcourus"],
            "litres_essence":  v["litres_essence"],
            "frais_voyage":    round(frais_total, 2),
            "nb_visites":      v["nb_visites"],
            "marge_estimee":   marge_estimee,
            "rentabilite_nette": rentabilite,
        })
    return faits

# ══════════════════════════════════════════════════════════════════════════════
# ÉTAPE 3 — CHARGEMENT PostgreSQL
# ══════════════════════════════════════════════════════════════════════════════
DDL_POSTGRES = """
-- Schéma DWH en étoile
DROP TABLE IF EXISTS fait_analyse_representant CASCADE;
DROP TABLE IF EXISTS dim_temps    CASCADE;
DROP TABLE IF EXISTS dim_vendeur  CASCADE;
DROP TABLE IF EXISTS dim_client   CASCADE;
DROP TABLE IF EXISTS dim_produit  CASCADE;
DROP TABLE IF EXISTS dim_geo      CASCADE;

CREATE TABLE dim_temps (
    id_dim_temps   SERIAL PRIMARY KEY,
    date_complete  DATE        NOT NULL,
    annee          SMALLINT    NOT NULL,
    trimestre      SMALLINT    NOT NULL,
    mois           SMALLINT    NOT NULL,
    lib_mois       VARCHAR(20) NOT NULL,
    semaine        SMALLINT    NOT NULL,
    jour           SMALLINT    NOT NULL,
    lib_jour       VARCHAR(20) NOT NULL,
    est_weekend    BOOLEAN     NOT NULL DEFAULT FALSE
);

CREATE TABLE dim_vendeur (
    id_dim_vendeur  SERIAL PRIMARY KEY,
    id_source       INT,
    nom             VARCHAR(50),
    prenom          VARCHAR(50),
    nom_complet     VARCHAR(100),
    salaire         NUMERIC(10,2),
    date_embauche   DATE,
    ville_base      VARCHAR(50),
    groupe          VARCHAR(30)
);

CREATE TABLE dim_client (
    id_dim_client     SERIAL PRIMARY KEY,
    id_source         INT,
    nom               VARCHAR(50),
    prenom            VARCHAR(50),
    adresse           VARCHAR(150),
    ville             VARCHAR(50),
    province          VARCHAR(60),
    pays              VARCHAR(30),
    personne_ressource VARCHAR(100),
    telephone         VARCHAR(20),
    email             VARCHAR(100),
    segment           VARCHAR(30)
);

CREATE TABLE dim_produit (
    id_dim_produit  INT PRIMARY KEY,
    nom             VARCHAR(100),
    categorie       VARCHAR(50),
    groupe          VARCHAR(30),
    prix_unitaire   NUMERIC(10,2),
    stock           INT,
    fournisseur     VARCHAR(60),
    date_lancement  DATE,
    actif           BOOLEAN DEFAULT TRUE
);

CREATE TABLE dim_geo (
    id_dim_geo  SERIAL PRIMARY KEY,
    ville       VARCHAR(50),
    province    VARCHAR(60),
    pays        VARCHAR(30)
);

CREATE TABLE fait_analyse_representant (
    id_fait             SERIAL PRIMARY KEY,
    id_dim_temps        INT NOT NULL REFERENCES dim_temps(id_dim_temps),
    id_dim_vendeur      INT NOT NULL REFERENCES dim_vendeur(id_dim_vendeur),
    id_dim_client       INT NOT NULL REFERENCES dim_client(id_dim_client),
    id_dim_geo          INT NOT NULL REFERENCES dim_geo(id_dim_geo),
    id_dim_produit      INT NOT NULL REFERENCES dim_produit(id_dim_produit),
    quantite_vendue     INT,
    montant_vente       NUMERIC(12,2),
    montant_precom      NUMERIC(12,2),
    km_parcourus        INT,
    litres_essence      NUMERIC(8,2),
    frais_voyage        NUMERIC(8,2),
    nb_visites          INT,
    marge_estimee       NUMERIC(12,2),
    rentabilite_nette   NUMERIC(12,2)
);

CREATE INDEX idx_fait_temps    ON fait_analyse_representant(id_dim_temps);
CREATE INDEX idx_fait_vendeur  ON fait_analyse_representant(id_dim_vendeur);
CREATE INDEX idx_fait_client   ON fait_analyse_representant(id_dim_client);
CREATE INDEX idx_fait_produit  ON fait_analyse_representant(id_dim_produit);
CREATE INDEX idx_fait_geo      ON fait_analyse_representant(id_dim_geo);
"""

def generate_load_sql(dim_temps, dim_vendeur, dim_client, dim_produit, dim_geo, faits) -> str:
    """Génère le script SQL complet de chargement PostgreSQL."""
    lines = [DDL_POSTGRES]
    lines.append("\n-- ════ DIM_TEMPS ════")
    for r in dim_temps:
        lines.append(
            f"INSERT INTO dim_temps(id_dim_temps,date_complete,annee,trimestre,mois,lib_mois,semaine,jour,lib_jour,est_weekend) "
            f"VALUES ({r['id_dim_temps']},'{r['date_complete']}',{r['annee']},{r['trimestre']},"
            f"{r['mois']},'{r['lib_mois']}',{r['semaine']},{r['jour']},'{r['lib_jour']}',{str(r['est_weekend']).upper()});"
        )
    lines.append("\n-- ════ DIM_VENDEUR ════")
    for r in dim_vendeur:
        lines.append(
            f"INSERT INTO dim_vendeur(id_dim_vendeur,id_source,nom,prenom,nom_complet,salaire,date_embauche,ville_base,groupe) "
            f"VALUES ({r['id_dim_vendeur']},{r['id_source']},'{r['nom']}','{r['prenom']}',"
            f"'{r['nom_complet']}',{r['salaire']},'{r['date_embauche']}','{r['ville_base']}','{r['groupe']}');"
        )
    lines.append("\n-- ════ DIM_CLIENT ════")
    for r in dim_client:
        addr     = r['adresse'].replace("'", "''")
        pr       = r['personne_ressource'].replace("'", "''")
        province = r['province'].replace("'", "''")
        nom      = r['nom'].replace("'", "''")
        prenom   = r['prenom'].replace("'", "''")
        lines.append(
            f"INSERT INTO dim_client(id_dim_client,id_source,nom,prenom,adresse,ville,province,pays,personne_ressource,telephone,email,segment) "
            f"VALUES ({r['id_dim_client']},{r['id_source']},'{nom}','{prenom}',"
            f"'{addr}','{r['ville']}','{province}','{r['pays']}','{pr}',"
            f"'{r['telephone']}','{r['email']}','{r['segment']}');"
        )
    lines.append("\n-- ════ DIM_PRODUIT ════")
    for r in dim_produit:
        dl = f"'{r['date_lancement']}'" if r['date_lancement'] else "NULL"
        # AFTER
        nom        = r['nom'].replace("'", "''")
        categorie  = r['categorie'].replace("'", "''")
        groupe     = r['groupe'].replace("'", "''")
        fournisseur= r['fournisseur'].replace("'", "''")
        lines.append(
            f"INSERT INTO dim_produit(id_dim_produit,nom,categorie,groupe,prix_unitaire,stock,fournisseur,date_lancement,actif) "
            f"VALUES ({r['id_dim_produit']},'{nom}','{categorie}','{groupe}',"
            f"{r['prix_unitaire']},{r['stock']},'{fournisseur}',{dl},{str(r['actif']).upper()});"
        )
    lines.append("\n-- ════ DIM_GEO ════")
    for r in dim_geo:
        lines.append(
            f"INSERT INTO dim_geo(id_dim_geo,ville,province,pays) "
            f"VALUES ({r['id_dim_geo']},'{r['ville']}','{r['province']}','{r['pays']}');"
        )
    lines.append("\n-- ════ FAIT_ANALYSE_REPRESENTANT ════")
    for f in faits:
        lines.append(
            f"INSERT INTO fait_analyse_representant("
            f"id_dim_temps,id_dim_vendeur,id_dim_client,id_dim_geo,id_dim_produit,"
            f"quantite_vendue,montant_vente,montant_precom,km_parcourus,litres_essence,"
            f"frais_voyage,nb_visites,marge_estimee,rentabilite_nette) "
            f"VALUES ({f['id_dim_temps']},{f['id_dim_vendeur']},{f['id_dim_client']},"
            f"{f['id_dim_geo']},{f['id_dim_produit']},"
            f"{f['quantite_vendue']},{f['montant_vente']},{f['montant_precom']},"
            f"{f['km_parcourus']},{f['litres_essence']},{f['frais_voyage']},"
            f"{f['nb_visites']},{f['marge_estimee']},{f['rentabilite_nette']});"
        )

    # Requêtes analytiques de validation
    lines.append("""
-- ════════════════════════════════════════════
-- REQUÊTES ANALYTIQUES DE VALIDATION
-- ════════════════════════════════════════════

-- 1. Ventes par vendeur (performance)
SELECT v.nom_complet, COUNT(*) AS nb_ventes,
       SUM(f.montant_vente) AS ca_total,
       AVG(f.montant_vente) AS ca_moyen,
       SUM(f.km_parcourus)  AS km_total
FROM fait_analyse_representant f
JOIN dim_vendeur v ON f.id_dim_vendeur = v.id_dim_vendeur
GROUP BY v.nom_complet ORDER BY ca_total DESC;

-- 2. Ventes par produit et catégorie
SELECT p.categorie, p.nom, SUM(f.quantite_vendue) AS qte,
       SUM(f.montant_vente) AS ca, SUM(f.marge_estimee) AS marge
FROM fait_analyse_representant f
JOIN dim_produit p ON f.id_dim_produit = p.id_dim_produit
GROUP BY p.categorie, p.nom ORDER BY ca DESC;

-- 3. Évolution mensuelle du CA
SELECT t.annee, t.lib_mois, t.mois,
       SUM(f.montant_vente) AS ca, SUM(f.frais_voyage) AS frais,
       SUM(f.rentabilite_nette) AS rentabilite
FROM fait_analyse_representant f
JOIN dim_temps t ON f.id_dim_temps = t.id_dim_temps
GROUP BY t.annee, t.lib_mois, t.mois ORDER BY t.annee, t.mois;

-- 4. Couverture géographique des vendeurs
SELECT v.nom_complet, g.province, COUNT(*) AS nb_visites,
       SUM(f.km_parcourus) AS km
FROM fait_analyse_representant f
JOIN dim_vendeur v ON f.id_dim_vendeur = v.id_dim_vendeur
JOIN dim_geo g     ON f.id_dim_geo     = g.id_dim_geo
GROUP BY v.nom_complet, g.province ORDER BY v.nom_complet, nb_visites DESC;

-- 5. Rentabilité par segment client
SELECT c.segment, COUNT(DISTINCT f.id_dim_client) AS nb_clients,
       SUM(f.montant_vente) AS ca, SUM(f.rentabilite_nette) AS rentabilite
FROM fait_analyse_representant f
JOIN dim_client c ON f.id_dim_client = c.id_dim_client
GROUP BY c.segment ORDER BY rentabilite DESC;
""")
    return "\n".join(lines)

# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def run_etl():
    print(f"\n{B}{'═'*60}{E}")
    print(f"{B}  ETL DataWarehouse — Imprimantes{E}")
    print(f"{B}{'═'*60}{E}\n")

    # ── E : Extraction ──────────────────────────────────────────────
    print(f"{Y}[E] EXTRACTION{E}")
    info("MySQL SQL → vendeurs + ventes")
    vendeurs, ventes = extract_mysql_sql(SQL_FILE)
    ok(f"{len(vendeurs)} vendeurs | {len(ventes)} ventes")

    info("Fichier TXT → clients")
    clients = extract_txt(TXT_FILE)
    ok(f"{len(clients)} clients")

    info("Excel → produits")
    produits = extract_excel(EXCEL_FILE)
    ok(f"{len(produits)} produits")

    info("JSON → feuilles de route")
    feuilles = extract_json(JSON_FILE)
    ok(f"{len(feuilles)} feuilles de route")

    # ── T : Transformation ──────────────────────────────────────────
    print(f"\n{Y}[T] TRANSFORMATION{E}")
    info("Dimension TEMPS — extraction date + attributs calendaires")
    dim_temps = transform_dim_temps(ventes, feuilles)
    ok(f"{len(dim_temps)} dates uniques")

    info("Dimension VENDEUR — clé surrogate + normalisation")
    dim_vendeur = transform_dim_vendeur(vendeurs)
    ok(f"{len(dim_vendeur)} vendeurs")

    info("Dimension CLIENT — nettoyage TXT + enrichissement")
    dim_client = transform_dim_client(clients)
    ok(f"{len(dim_client)} clients")

    info("Dimension PRODUIT — conversion types Excel")
    dim_produit = transform_dim_produit(produits)
    ok(f"{len(dim_produit)} produits")

    info("Dimension GEO — extraction villes + mapping province")
    dim_geo = transform_dim_geo(ventes, feuilles)
    ok(f"{len(dim_geo)} zones géographiques")

    info("Table de FAITS — résolution clés + calcul mesures")
    faits = transform_fait(ventes, feuilles, dim_temps, dim_vendeur,
                           dim_client, dim_produit, dim_geo)
    ok(f"{len(faits)} lignes de faits chargées")
    rejects = len(ventes) - len(faits)
    if rejects > 0:
        warn(f"{rejects} lignes rejetées (clés orphelines)")

    # ── L : Load ────────────────────────────────────────────────────
    print(f"\n{Y}[L] CHARGEMENT{E}")
    info("Génération du script SQL PostgreSQL")
    sql_load = generate_load_sql(dim_temps, dim_vendeur, dim_client,
                                 dim_produit, dim_geo, faits)
    out_sql = Path(__file__).parent.parent / "Dtw1" / "05_load_postgres.sql"
    out_sql.write_text("SET client_encoding = 'UTF8';\n" + sql_load, encoding="utf-8")
    ok(f"Script généré → {out_sql.name}")

    # ── Rapport ─────────────────────────────────────────────────────
    print(f"\n{B}{'═'*60}{E}")
    print(f"{G}  ✅ ETL TERMINÉ AVEC SUCCÈS{E}")
    print(f"{B}{'═'*60}{E}")
    print(f"""
  📊 Résumé du chargement :
     DIM_TEMPS      : {len(dim_temps):>5} lignes
     DIM_VENDEUR    : {len(dim_vendeur):>5} lignes
     DIM_CLIENT     : {len(dim_client):>5} lignes
     DIM_PRODUIT    : {len(dim_produit):>5} lignes
     DIM_GEO        : {len(dim_geo):>5} lignes
     FAIT           : {len(faits):>5} lignes
     Rejetées       : {rejects:>5} lignes

  📁 Fichier de chargement : sources/05_load_postgres.sql
""")
    return True

if __name__ == "__main__":
    success = run_etl()
    sys.exit(0 if success else 1)

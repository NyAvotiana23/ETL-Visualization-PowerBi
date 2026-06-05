#!/usr/bin/env python3
"""
load_staging.py
---------------
Lit les 4 fichiers sources (SQL, TXT, XLSX, JSON) et génère un script
PostgreSQL de staging prêt à exécuter.

Toutes les colonnes sont définies en TEXT ou VARCHAR pour éviter tout
conflit de types sur des données brutes non nettoyées.

Sources attendues (chemins modifiables ci-dessous) :
    01_source_mysql_printer_sales.sql
    02_route_logs.txt
    03_sales_promises.xlsx
    04_fuel_expenses.json

Sortie :
    staging_load.sql   — script PostgreSQL exécutable directement avec psql
"""

import re
import json
import os
import sys
from pathlib import Path
from datetime import datetime

# ── Dépendances optionnelles ──────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    sys.exit("Erreur : 'openpyxl' est requis. Installez-le avec : pip install openpyxl")

# ── Chemins des fichiers sources ──────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
SQL_FILE   = BASE_DIR / "01_source_mysql_printer_sales.sql"
TXT_FILE   = BASE_DIR / "02_route_logs.txt"
XLSX_FILE  = BASE_DIR / "03_sales_promises.xlsx"
JSON_FILE  = BASE_DIR / "04_fuel_expenses.json"
OUTPUT_SQL = BASE_DIR / "staging_load.sql"

# ── Schéma cible ──────────────────────────────────────────────────────────────
STAGING_SCHEMA = "staging"

# =============================================================================
# Utilitaires SQL
# =============================================================================

def esc(val) -> str:
    """Échappe une valeur pour une chaîne littérale PostgreSQL (simple quote)."""
    if val is None:
        return "NULL"
    s = str(val).replace("'", "''")
    return f"'{s}'"


def dict_to_insert(schema: str, table: str, row: dict) -> str:
    """Génère un INSERT INTO … VALUES … à partir d'un dict."""
    cols = ", ".join(f'"{k}"' for k in row.keys())
    vals = ", ".join(esc(v) for v in row.values())
    return f'INSERT INTO {schema}."{table}" ({cols}) VALUES ({vals});'


# =============================================================================
# 1. Lecture du fichier SQL MySQL
#    Tables : hr_vendeurs, ref_clients, ref_produits, sales_orders
# =============================================================================

# Colonnes déclarées dans le DDL MySQL → utilisées pour mapper les INSERT
MYSQL_TABLES = {
    "hr_vendeurs": [
        "seller_code", "first_name", "last_name", "email", "salary",
        "hire_date", "home_country", "home_region", "home_city", "manager_code",
    ],
    "ref_clients": [
        "customer_code", "customer_name", "sector", "country",
        "region_name", "city", "postal_code", "created_at",
    ],
    "ref_produits": [
        "product_code", "product_name", "category_name", "range_name",
        "list_price", "active_flag", "launch_date",
    ],
    "sales_orders": [
        "order_id", "order_date", "seller_code", "customer_code",
        "product_code", "quantity", "unit_price", "discount_pct",
        "order_status", "promised_delivery_date", "created_at",
    ],
}

# Ordre réel des colonnes dans chaque INSERT du fichier SQL (suit l'ordre
# déclaré dans VALUES, qui correspond à l'ordre des colonnes de l'INSERT)
MYSQL_INSERT_COL_ORDER = {
    "hr_vendeurs":  MYSQL_TABLES["hr_vendeurs"],
    "ref_clients":  ["customer_code", "customer_name", "sector", "country",
                     "city", "postal_code", "created_at", "region_name"],
    "ref_produits": ["product_code", "product_name", "range_name",
                     "list_price", "active_flag", "launch_date", "category_name"],
    "sales_orders": MYSQL_TABLES["sales_orders"],
}


def parse_mysql_values(values_str: str) -> list[str]:
    """
    Décompose la partie VALUES d'un INSERT MySQL en liste de valeurs,
    en gérant correctement les virgules à l'intérieur des chaînes.
    """
    vals = []
    current = ""
    in_quote = False
    i = 0
    while i < len(values_str):
        ch = values_str[i]
        if ch == "'" and not in_quote:
            in_quote = True
            current += ch
        elif ch == "'" and in_quote:
            # double quote d'échappement MySQL ''
            if i + 1 < len(values_str) and values_str[i + 1] == "'":
                current += "''"
                i += 2
                continue
            in_quote = False
            current += ch
        elif ch == "," and not in_quote:
            vals.append(current.strip())
            current = ""
        else:
            current += ch
        i += 1
    if current.strip():
        vals.append(current.strip())
    return vals


def clean_mysql_value(raw: str):
    """Convertit une valeur MySQL brute en valeur Python."""
    raw = raw.strip()
    if raw.upper() == "NULL":
        return None
    if raw.startswith("'") and raw.endswith("'"):
        # Enlever les quotes externes et dé-doubler les '' internes
        return raw[1:-1].replace("''", "'")
    return raw  # nombre ou autre littéral


def load_mysql_sql(path: Path) -> dict[str, list[dict]]:
    """
    Parse le fichier SQL MySQL et retourne un dict
    { nom_table : [{ col: val, … }, …] }
    """
    data: dict[str, list[dict]] = {t: [] for t in MYSQL_TABLES}

    # Regex pour capturer le nom de table et les valeurs d'un INSERT
    insert_re = re.compile(
        r"INSERT INTO\s+(\w+)\s*\([^)]+\)\s*VALUES\s*\((.+?)\)\s*;",
        re.IGNORECASE | re.DOTALL,
    )

    text = path.read_text(encoding="utf-8", errors="replace")

    for m in insert_re.finditer(text):
        table = m.group(1).lower()
        if table not in MYSQL_TABLES:
            continue
        raw_vals = parse_mysql_values(m.group(2))
        col_order = MYSQL_INSERT_COL_ORDER[table]
        # Si le nombre de valeurs ne correspond pas, on tronque/complète
        row = {}
        for i, col in enumerate(col_order):
            row[col] = clean_mysql_value(raw_vals[i]) if i < len(raw_vals) else None
        # Pour ref_clients on remet les colonnes dans l'ordre canonique
        if table == "ref_clients":
            ordered_row = {c: row.get(c) for c in MYSQL_TABLES["ref_clients"]}
            data[table].append(ordered_row)
        else:
            data[table].append(row)

    return data


# =============================================================================
# 2. Lecture du fichier TXT (pipe-delimited) : route_logs
# =============================================================================

def load_txt_pipe(path: Path) -> list[dict]:
    """Lit un fichier TSV/PSV avec séparateur | et retourne une liste de dicts."""
    rows = []
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    if not lines:
        return rows

    header = [h.strip() for h in lines[0].split("|")]

    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split("|")]
        # Compléter si nombre de colonnes insuffisant
        while len(parts) < len(header):
            parts.append("")
        row = {header[i]: parts[i] for i in range(len(header))}
        rows.append(row)

    return rows


# =============================================================================
# 3. Lecture du fichier XLSX : sales_promises (feuille "promesses_vente")
# =============================================================================

def load_xlsx(path: Path, sheet_name: str = "promesses_vente") -> list[dict]:
    """Lit une feuille XLSX et retourne une liste de dicts."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Recherche de la feuille (insensible à la casse)
    target = None
    for name in wb.sheetnames:
        if name.lower() == sheet_name.lower():
            target = name
            break
    if target is None:
        # Fallback : première feuille
        target = wb.sheetnames[0]
        print(f"[WARN] Feuille '{sheet_name}' introuvable, utilisation de '{target}'.")

    ws = wb[target]
    rows_iter = ws.iter_rows(values_only=True)

    header = None
    data = []
    for row in rows_iter:
        values = [str(v).strip() if v is not None else "" for v in row]
        if header is None:
            header = values
            continue
        if not any(values):  # ligne vide → skip
            continue
        row_dict = {header[i]: values[i] for i in range(min(len(header), len(values)))}
        data.append(row_dict)

    wb.close()
    return data


# =============================================================================
# 4. Lecture du fichier JSON : fuel_expenses
# =============================================================================

def load_json(path: Path) -> list[dict]:
    """
    Charge le fichier JSON. Accepte un tableau de dicts ou un dict racine
    contenant une liste. Retourne toujours une liste de dicts aplatis.
    """
    raw = json.loads(path.read_text(encoding="utf-8", errors="replace"))

    if isinstance(raw, list):
        records = raw
    elif isinstance(raw, dict):
        # Cherche la première valeur qui est une liste
        for v in raw.values():
            if isinstance(v, list):
                records = v
                break
        else:
            # Traite le dict comme un seul enregistrement
            records = [raw]
    else:
        records = [{"value": raw}]

    # Aplatir les dicts imbriqués avec un préfixe
    flat_records = []
    for rec in records:
        flat = {}
        _flatten(rec, "", flat)
        flat_records.append(flat)
    return flat_records


def _flatten(obj, prefix: str, result: dict):
    """Aplatit récursivement un dict/list en un dict à un niveau."""
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{prefix}{k}" if prefix == "" else f"{prefix}__{k}"
            _flatten(v, new_key, result)
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            _flatten(v, f"{prefix}__{i}" if prefix else str(i), result)
    else:
        result[prefix] = None if obj is None else str(obj)


# =============================================================================
# 5. Génération du DDL de staging (tout en TEXT)
# =============================================================================

def cols_as_text(columns: list[str]) -> str:
    """Retourne la liste des colonnes DDL toutes en TEXT."""
    lines = [f'    "{c}" TEXT' for c in columns]
    return ",\n".join(lines)


def generate_create_table(schema: str, table: str, columns: list[str]) -> str:
    return (
        f'CREATE TABLE IF NOT EXISTS {schema}."{table}" (\n'
        f'{cols_as_text(columns)}\n'
        f');\n'
    )


def columns_from_rows(rows: list[dict]) -> list[str]:
    """Déduit les colonnes à partir de la liste de lignes (union de toutes les clés)."""
    seen = {}
    for r in rows:
        for k in r.keys():
            seen[k] = True
    return list(seen.keys())


# =============================================================================
# 6. Assemblage du script SQL de staging
# =============================================================================

def build_staging_sql(
    mysql_data: dict[str, list[dict]],
    route_logs: list[dict],
    sales_promises: list[dict],
    fuel_expenses: list[dict],
) -> str:

    lines = []

    # ── En-tête ───────────────────────────────────────────────────────────────
    lines.append("-- ============================================================")
    lines.append("-- staging_load.sql")
    lines.append(f"-- Généré le {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("-- Toutes les colonnes sont en TEXT pour éviter les conflits")
    lines.append("-- de types sur des données brutes non nettoyées.")
    lines.append("-- ============================================================")
    lines.append("")
    lines.append(f"CREATE SCHEMA IF NOT EXISTS {STAGING_SCHEMA};")
    lines.append("")

    # ── Helper interne ────────────────────────────────────────────────────────
    def section(title: str):
        lines.append("")
        lines.append("-- " + "─" * 60)
        lines.append(f"-- {title}")
        lines.append("-- " + "─" * 60)
        lines.append("")

    def dump_table(table_name: str, rows: list[dict], canonical_cols: list[str] | None = None):
        """Génère DROP + CREATE + TRUNCATE + INSERTs pour une table."""
        if not rows:
            lines.append(f"-- (aucune donnée pour {table_name})")
            return

        all_cols = canonical_cols if canonical_cols else columns_from_rows(rows)

        lines.append(f'DROP TABLE IF EXISTS {STAGING_SCHEMA}."{table_name}";')
        lines.append(generate_create_table(STAGING_SCHEMA, table_name, all_cols))

        count = 0
        for row in rows:
            # S'assurer que toutes les colonnes déclarées existent dans le row
            complete_row = {c: row.get(c) for c in all_cols}
            lines.append(dict_to_insert(STAGING_SCHEMA, table_name, complete_row))
            count += 1

        lines.append(f"-- {count} ligne(s) insérée(s) dans {table_name}")
        lines.append("")

    # ── 1. Tables MySQL ───────────────────────────────────────────────────────
    section("SOURCE 1 : MySQL — printer_sales")

    for table_name, canonical in MYSQL_TABLES.items():
        rows = mysql_data.get(table_name, [])
        lines.append(f"-- Table : {table_name} ({len(rows)} lignes)")
        dump_table(table_name, rows, canonical)

    # ── 2. Route Logs (TXT) ───────────────────────────────────────────────────
    section("SOURCE 2 : TXT pipe-delimited — route_logs")
    lines.append(f"-- Fichier : {TXT_FILE.name} ({len(route_logs)} lignes)")
    dump_table("stg_route_logs", route_logs)

    # ── 3. Sales Promises (XLSX) ──────────────────────────────────────────────
    section("SOURCE 3 : XLSX — sales_promises")
    lines.append(f"-- Fichier : {XLSX_FILE.name} ({len(sales_promises)} lignes)")
    dump_table("stg_sales_promises", sales_promises)

    # ── 4. Fuel Expenses (JSON) ───────────────────────────────────────────────
    section("SOURCE 4 : JSON — fuel_expenses")
    lines.append(f"-- Fichier : {JSON_FILE.name} ({len(fuel_expenses)} lignes)")
    dump_table("stg_fuel_expenses", fuel_expenses)

    # ── Pied de page ──────────────────────────────────────────────────────────
    lines.append("")
    lines.append("-- ============================================================")
    lines.append("-- FIN DU SCRIPT DE STAGING")
    lines.append("-- ============================================================")

    return "\n".join(lines)


# =============================================================================
# Point d'entrée
# =============================================================================

def main():
    print("=" * 60)
    print("  load_staging.py — Génération du script PostgreSQL de staging")
    print("=" * 60)

    # Vérification de l'existence des fichiers
    for f in [SQL_FILE, TXT_FILE, XLSX_FILE]:
        if not f.exists():
            sys.exit(f"[ERREUR] Fichier introuvable : {f}")

    # JSON optionnel (peut être vide ou absent)
    fuel_expenses: list[dict] = []
    if JSON_FILE.exists() and JSON_FILE.stat().st_size > 2:
        try:
            fuel_expenses = load_json(JSON_FILE)
            print(f"[OK] JSON        : {len(fuel_expenses)} enregistrement(s)")
        except Exception as e:
            print(f"[WARN] JSON vide ou invalide ({e}), table sera vide.")
    else:
        print("[INFO] JSON      : fichier vide ou absent — table vide.")

    # 1. MySQL SQL
    print(f"[...] Lecture de {SQL_FILE.name} …")
    mysql_data = load_mysql_sql(SQL_FILE)
    for t, rows in mysql_data.items():
        print(f"[OK]  MySQL      : {t} — {len(rows)} ligne(s)")

    # 2. Route Logs TXT
    print(f"[...] Lecture de {TXT_FILE.name} …")
    route_logs = load_txt_pipe(TXT_FILE)
    print(f"[OK]  TXT        : stg_route_logs — {len(route_logs)} ligne(s)")

    # 3. Sales Promises XLSX
    print(f"[...] Lecture de {XLSX_FILE.name} …")
    sales_promises = load_xlsx(XLSX_FILE)
    print(f"[OK]  XLSX       : stg_sales_promises — {len(sales_promises)} ligne(s)")

    # 4. Génération du SQL
    print("[...] Génération du script PostgreSQL …")
    sql_content = build_staging_sql(mysql_data, route_logs, sales_promises, fuel_expenses)

    OUTPUT_SQL.write_text(sql_content, encoding="utf-8")
    size_kb = OUTPUT_SQL.stat().st_size / 1024
    print(f"[OK]  Script généré : {OUTPUT_SQL.name}  ({size_kb:.1f} Ko)")
    print()

    # Résumé
    total_rows = (
        sum(len(r) for r in mysql_data.values())
        + len(route_logs)
        + len(sales_promises)
        + len(fuel_expenses)
    )
    print("─" * 60)
    print(f"  Lignes totales chargées : {total_rows}")
    print(f"  Tables de staging créées :")
    print(f"    • staging.hr_vendeurs          ({len(mysql_data['hr_vendeurs'])} lignes)")
    print(f"    • staging.ref_clients          ({len(mysql_data['ref_clients'])} lignes)")
    print(f"    • staging.ref_produits         ({len(mysql_data['ref_produits'])} lignes)")
    print(f"    • staging.sales_orders         ({len(mysql_data['sales_orders'])} lignes)")
    print(f"    • staging.stg_route_logs       ({len(route_logs)} lignes)")
    print(f"    • staging.stg_sales_promises   ({len(sales_promises)} lignes)")
    print(f"    • staging.stg_fuel_expenses    ({len(fuel_expenses)} lignes)")
    print("─" * 60)
    print(f"\n  Pour exécuter : psql -d <votre_db> -f {OUTPUT_SQL.name}")
    print()


if __name__ == "__main__":
    main()
